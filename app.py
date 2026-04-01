# app.py  ─  IMEMSA · Sistema de Compra de Motores Yamaha
# ══════════════════════════════════════════════════════════
# Ejecutar:  streamlit run app.py
# ══════════════════════════════════════════════════════════

from __future__ import annotations
import base64
import io
import os
from datetime import datetime, date, timedelta

import streamlit as st

st.set_page_config(
    page_title="IMEMSA · Motores Yamaha",
    page_icon="⚓",
    layout="wide",
    initial_sidebar_state="expanded",
)

from utils.auth         import verify_login
from utils.constants    import (USERS, ACTIVITIES_TEMPLATE, PHASES,
                                MOTOR_MODELS, STATUS_LABELS,
                                ORDER_STATUS_LABELS, MAX_ANNUAL_ORDERS,
                                IMEMSA_NAVY, IMEMSA_RED)
from utils.data_manager import (load_data, save_data, create_order,
                                request_closure, approve_closure,
                                get_orders_for_user, get_semaphore,
                                get_semaphore_summary, get_red_activities,
                                get_my_pending_activities)
from utils.email_utils  import send_activation_email, send_overdue_alert


# ══════════════════════════════════════════════════════════
#  CAPA DE DATOS EN MEMORIA
# ══════════════════════════════════════════════════════════

def _app_data() -> dict:
    if "_app_data_store" not in st.session_state:
        st.session_state["_app_data_store"] = load_data()
    return st.session_state["_app_data_store"]


def _app_save(data: dict) -> None:
    """Guarda en session_state (inmediato) y en Sheets. Avisa si Sheets falla."""
    st.session_state["_app_data_store"] = data
    try:
        from utils.sheets_manager import _gsheets_available, save_to_sheets, _strip_evidence
        if _gsheets_available():
            ok = save_to_sheets(data)
            if not ok:
                st.toast("⚠️ No se pudo guardar en Google Sheets. Intenta de nuevo.", icon="⚠️")
        else:
            save_data(data)
    except Exception as e:
        st.toast(f"⚠️ Error al guardar: {e}", icon="⚠️")


def _app_invalidate() -> None:
    st.session_state.pop("_app_data_store", None)


# ══════════════════════════════════════════════════════════
#  CSS GLOBAL
# ══════════════════════════════════════════════════════════

def inject_css() -> None:
    st.markdown("""
    <style>
    @import url('https://fonts.googleapis.com/css2?family=Barlow+Condensed:wght@400;600;700;800&family=Source+Sans+3:wght@300;400;600;700&display=swap');

    html, body, [data-testid="stAppViewContainer"] {
        font-family: 'Source Sans 3', sans-serif; background: #F0F3F9;
    }
    h1,h2,h3,h4 { font-family: 'Barlow Condensed', sans-serif; letter-spacing:.5px; }

    [data-testid="stSidebar"] {
        background: linear-gradient(180deg, #0D2B6E 0%, #091D4E 100%) !important;
        border-right: none !important;
    }
    [data-testid="stSidebar"] * { color: #E8EDF7 !important; }
    [data-testid="stSidebar"] .sidebar-divider {
        border-top: 1px solid rgba(255,255,255,.15); margin: 12px 0;
    }
    [data-testid="stSidebar"] [data-testid="stButton"] button {
        background: rgba(255,255,255,.08) !important;
        border: 1px solid rgba(255,255,255,.18) !important;
        color: #fff !important; border-radius: 8px !important;
        font-family: 'Source Sans 3', sans-serif !important;
        font-weight: 600 !important; transition: background .2s;
    }
    [data-testid="stSidebar"] [data-testid="stButton"] button:hover {
        background: rgba(255,255,255,.18) !important;
    }

    .main .block-container { padding: 1.5rem 2rem 3rem 2rem; max-width:1400px; }

    .metric-card {
        background: #fff; border-radius: 12px; padding: 20px 24px;
        box-shadow: 0 1px 6px rgba(13,43,110,.10);
        border-left: 4px solid var(--card-accent, #0D2B6E);
        transition: transform .15s, box-shadow .15s;
    }
    .metric-card:hover { transform: translateY(-2px); box-shadow: 0 4px 14px rgba(13,43,110,.14); }
    .metric-card .mc-value {
        font-family: 'Barlow Condensed', sans-serif;
        font-size: 2.4rem; font-weight: 800; line-height: 1; color: #0D2B6E;
    }
    .metric-card .mc-label {
        font-size: .82rem; font-weight: 600; text-transform: uppercase;
        letter-spacing: .6px; color: #8592A3; margin-top: 4px;
    }

    .order-card {
        background: #fff; border-radius: 12px; padding: 20px 22px;
        box-shadow: 0 1px 6px rgba(13,43,110,.09); margin-bottom: 14px;
        border-top: 3px solid #0D2B6E; transition: transform .15s, box-shadow .15s;
    }
    .order-card:hover { transform: translateY(-2px); box-shadow: 0 4px 16px rgba(13,43,110,.14); }
    .order-number {
        font-family: 'Barlow Condensed', sans-serif;
        font-size: 1.15rem; font-weight: 700; color: #0D2B6E;
    }
    .order-model { font-size: .85rem; color: #4B5563; margin-top: 2px; }
    .badge {
        display: inline-block; padding: 3px 10px; border-radius: 20px;
        font-size: .72rem; font-weight: 700; text-transform: uppercase; letter-spacing: .5px;
    }
    .badge-active    { background: #DBEAFE; color: #1E40AF; }
    .badge-completed { background: #D1FAE5; color: #065F46; }
    .badge-cancelled { background: #F3F4F6; color: #6B7280; }

    .progress-wrap { background: #E5E7EB; border-radius: 6px; height: 8px; overflow:hidden; margin: 8px 0; }
    .progress-fill { height: 8px; border-radius: 6px;
                     background: linear-gradient(90deg, #0D2B6E, #2563EB); transition: width .4s; }

    .semaphore-dot {
        display: inline-block; width: 12px; height: 12px;
        border-radius: 50%; margin-right: 6px; vertical-align: middle;
    }
    .sem-green  { background: #22C55E; box-shadow: 0 0 6px #22C55E88; }
    .sem-yellow { background: #F59E0B; box-shadow: 0 0 6px #F59E0B88; }
    .sem-red    { background: #EF4444; box-shadow: 0 0 6px #EF444488; animation: pulse-red 1.2s infinite; }
    .sem-gray   { background: #D1D5DB; }
    @keyframes pulse-red {
        0%,100% { box-shadow: 0 0 6px #EF444488; }
        50%      { box-shadow: 0 0 14px #EF4444CC; }
    }

    .act-row {
        background: #fff; border-radius: 10px; padding: 14px 18px; margin-bottom: 8px;
        border-left: 4px solid var(--act-color, #D1D5DB);
        box-shadow: 0 1px 4px rgba(0,0,0,.06);
    }
    .act-row-completed  { --act-color: #22C55E; background: #F0FDF4; }
    .act-row-inprogress { --act-color: #2563EB; }
    .act-row-waiting    { --act-color: #F59E0B; background: #FFFBEB; }
    .act-row-pending    { --act-color: #D1D5DB; }
    .act-row-blocked    { --act-color: #EF4444; background: #FEF2F2; }
    .act-name { font-family:'Barlow Condensed',sans-serif; font-size:1.05rem; font-weight:700; color:#0D2B6E; }
    .act-meta { font-size:.78rem; color:#8592A3; margin-top:2px; }

    .phase-chip {
        display: inline-block; padding: 2px 10px; border-radius: 12px;
        font-size: .70rem; font-weight: 700; text-transform: uppercase;
        letter-spacing: .5px; background: #EEF2FF; color: #3730A3; margin-right: 6px;
    }

    .section-header {
        font-family: 'Barlow Condensed', sans-serif;
        font-size: 1.5rem; font-weight: 800; color: #0D2B6E;
        border-bottom: 3px solid #C41E2E;
        padding-bottom: 6px; margin-bottom: 20px; letter-spacing: .4px;
    }

    .avatar {
        display: inline-flex; align-items:center; justify-content:center;
        width: 38px; height: 38px; border-radius: 50%;
        font-family: 'Barlow Condensed', sans-serif;
        font-weight: 800; font-size: 1rem; color: #fff; margin-right: 8px;
    }

    .nav-active button {
        background: rgba(196,30,46,.20) !important;
        border-color: #C41E2E !important; color: #fff !important;
    }

    [data-testid="stTextInput"] input,
    [data-testid="stSelectbox"] select,
    [data-testid="stNumberInput"] input,
    [data-testid="stTextArea"] textarea {
        border-radius: 8px !important;
        border: 1.5px solid #D1D9E8 !important;
        font-family: 'Source Sans 3', sans-serif !important;
        background: #FFFFFF !important; color: #1F2937 !important;
    }
    [data-testid="stTextInput"] input::placeholder,
    [data-testid="stTextArea"] textarea::placeholder { color: #9CA3AF !important; }

    /* Selectbox — texto negro en el campo seleccionado */
    [data-testid="stSelectbox"] > div > div {
        background: #FFFFFF !important;
        border: 1.5px solid #D1D9E8 !important;
        border-radius: 8px !important;
    }
    [data-testid="stSelectbox"] span,
    [data-testid="stSelectbox"] p,
    [data-testid="stSelectbox"] div[data-baseweb="select"] span {
        color: #1F2937 !important;
    }
    /* Labels de selectbox y widgets fuera del sidebar */
    .main [data-testid="stWidgetLabel"] p,
    .main [data-testid="stWidgetLabel"] span,
    [data-testid="stWidgetLabel"] p,
    [data-testid="stWidgetLabel"] span,
    [data-testid="stForm"] label,
    .main label p, .main label span { color: #1F2937 !important; }

    /* Tabs — texto oscuro (no heredar blanco del sidebar) */
    [data-testid="stTabs"] [data-baseweb="tab"] p,
    [data-testid="stTabs"] [data-baseweb="tab"] span,
    [data-testid="stTabs"] button[role="tab"] p,
    [data-testid="stTabs"] button[role="tab"] span,
    [data-testid="stTabs"] button[role="tab"] { color: #0D2B6E !important; }
    [data-testid="stTabs"] button[role="tab"][aria-selected="true"] {
        color: #0D2B6E !important; font-weight: 700 !important;
        border-bottom: 3px solid #C41E2E !important;
    }

    [data-testid="stSidebar"] p,
    [data-testid="stSidebar"] span,
    [data-testid="stSidebar"] div,
    [data-testid="stSidebar"] label { color: #E8EDF7 !important; }

    [data-testid="stButton"] button[kind="primary"] {
        background: #0D2B6E !important; border: none !important;
        border-radius: 8px !important; font-family: 'Source Sans 3', sans-serif !important;
        font-weight: 700 !important; letter-spacing: .3px !important;
        transition: background .2s !important;
    }
    [data-testid="stButton"] button[kind="primary"]:hover { background: #C41E2E !important; }
    .stAlert { border-radius: 8px !important; }

    [data-testid="stExpander"] summary p,
    [data-testid="stExpander"] summary span,
    [data-testid="stExpander"] details summary p {
        color: #0D2B6E !important; font-weight: 600 !important;
    }
    [data-testid="stExpander"] {
        background: #EEF2FF !important;
        border: 1.5px solid #C7D2FE !important; border-radius: 8px !important;
    }
    </style>
    """, unsafe_allow_html=True)


# ══════════════════════════════════════════════════════════
#  LOGO IMEMSA
# ══════════════════════════════════════════════════════════

def _load_logo_b64() -> str:
    logo_path = os.path.join(os.path.dirname(__file__), "assets", "logo_imemsa.png")
    if os.path.exists(logo_path):
        with open(logo_path, "rb") as f:
            return base64.b64encode(f.read()).decode()
    return ""

_LOGO_B64 = _load_logo_b64()

def _logo_img_tag(width: int = 160, style: str = "") -> str:
    if _LOGO_B64:
        return (f'<img src="data:image/png;base64,{_LOGO_B64}" '
                f'width="{width}" style="display:block;{style}" alt="IMEMSA"/>')
    return '<span style="font-family:Barlow Condensed,sans-serif;font-size:1.6rem;font-weight:800;color:#fff;">IMEMSA</span>'


def logo_sidebar() -> None:
    st.markdown(
        '<div style="padding:16px 16px 10px 16px;">'
        '<img src="data:image/png;base64,/9j/4AAQSkZJRgABAQAAAQABAAD/4gHYSUNDX1BST0ZJTEUAAQEAAAHIAAAAAAQwAABtbnRyUkdCIFhZWiAH4AABAAEAAAAAAABhY3NwAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAQAA9tYAAQAAAADTLQAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAlkZXNjAAAA8AAAACRyWFlaAAABFAAAABRnWFlaAAABKAAAABRiWFlaAAABPAAAABR3dHB0AAABUAAAABRyVFJDAAABZAAAAChnVFJDAAABZAAAAChiVFJDAAABZAAAAChjcHJ0AAABjAAAADxtbHVjAAAAAAAAAAEAAAAMZW5VUwAAAAgAAAAcAHMAUgBHAEJYWVogAAAAAAAAb6IAADj1AAADkFhZWiAAAAAAAABimQAAt4UAABjaWFlaIAAAAAAAACSgAAAPhAAAts9YWVogAAAAAAAA9tYAAQAAAADTLXBhcmEAAAAAAAQAAAACZmYAAPKnAAANWQAAE9AAAApbAAAAAAAAAABtbHVjAAAAAAAAAAEAAAAMZW5VUwAAACAAAAAcAEcAbwBvAGcAbABlACAASQBuAGMALgAgADIAMAAxADb/2wBDAAUDBAQEAwUEBAQFBQUGBwwIBwcHBw8LCwkMEQ8SEhEPERETFhwXExQaFRERGCEYGh0dHx8fExciJCIeJBweHx7/2wBDAQUFBQcGBw4ICA4eFBEUHh4eHh4eHh4eHh4eHh4eHh4eHh4eHh4eHh4eHh4eHh4eHh4eHh4eHh4eHh4eHh4eHh7/wAARCAA8AVIDASIAAhEBAxEB/8QAHAAAAwADAQEBAAAAAAAAAAAAAAUGAwQHAQII/8QAQRAAAQQBAQQFBwkGBwEAAAAAAQACAwQRBQYSITETIkFRkRQVYXGBktEyQlOTobHB4fAHI1JUcvEWMzRDRGKCNf/EABsBAQACAwEBAAAAAAAAAAAAAAABAgMFBgQH/8QAMhEAAgIBAgUBBAkFAAAAAAAAAAECEQMEBRIhMUGRBgcTUWEUFhciU1RxgdEyQqGx8P/aAAwDAQACEQMRAD8A/MSELquz2lU26PWElWJzzGC4lgJJPFdRum5w2+EZyjdujJ6Z9NZd/wA08WOajwq7fPvRypGCuzDTaI/4kH1TfgvfN9L+Ur/VN+C0n1txfhvyjsvsp1X5iPhnGELs/m6l/KV/qm/BeebaOf8ASw/Vt+CfW3F+G/KH2U6n8xHwzjKF03a8afp2jzPFeDpJG7kY6MZyVEbPVmyvkke0ENwBkZW+2zX/AE/G8ijwq6OI9R7C9j1K00sqnKrdLoKUKquabWmYf3TWuxzaMKZsxOgnfE7m04Wyao54xoVTptOJtOMPjYXYyctBWc065/2me6FPACPQtzVoBBcc0DDTxC1Y2OkeGMGXE4AVaB8oVPpumxQwDpGNc88SSAVtGnXPOJnuj4K3ACOQmm0AijmZFGxrcDJwAEz0ulCKcZfG0uIyctBUKNugTCFYeR1/ome6EeR1/ome6FPACPQrHyOv9Ez3R8Fq6jHVr1nPMbM9nVCcAJhC3tFgFi8N5uWtGTwVF5HX+iZ7oUKNgj0Kl1WjEaTzHG0OaMjDQFNKGqAIW/pFB1qUPeD0QPH0qhbTrgY6JnuhSo2CPQq2xXrRROkMTOqM/JHYpmtBJbn3IxzOT3BHGgYF6GuPJpPsVRS0uCBg3mB7+0kZW4ImYxjwOFKgCLII5gheK0kgikGHsDh6RlK7+jRuaXwdQ93YjgwT6F9SNLHljuYOCvlUAIWelXfZnEbB6z3Knio12MDRGw4Ha0KyjYJFCr5acDmECNmSP4Qlex4jbtPBFNGx7XFzMOGRnBWHUT9xilkq6Tfg9Wi030rU48F1xNK/hboSIXWr2z+mW4S19OJrscHNbukeC5nrlA6bqctQu3gw9U94Ws23ecO4NxgmpLszo/Ufo7WbDCOXLJSg3Vr4/NM0UIQtuckA5rpU+09HTNOrRvDpJjCwhjOzh2nsXNV1j9iui6DNptrW9VEE00Em6BNgtjGOeDwWt3Hb8OsUXluo9vidBsfqLVbKsr0tcU0lb51T7L+Sf/x4Qf8A5x3e/pOP3LPDt3Td/nVJ2f0kOXS7m3Gw1d/Qvs1ZMcOpFvAeAXjNQ/ZzrWGvGjyuPY+JrXfcCvBLZNC1zxV+7Nnj9fb9B37+/wBYx/ggYttNFcOu+dh7jDn8VhubcacxjvJ4p5n9mW7g+9dBm/Z3sPqDS+vVaze7YLB+JWBv7J9k4nCR3ljmjiQ6fh9gWGOw7cnbT/Sz15PaNvkocKnFfNRVnENb1a3q1npbDsNHBjByaE22ei6Oi155vJd+vBY9uzp8m1U1TSImR1ISIYgzkccCfTxTKuxscLWN5AYHqC6bT44Y4KMFSXY4jU6nLqcssuaTlJ9W+59vzunvKlpB5XrJaDkOlwD6B/ZUeoSiGpJJ3AlIdnoukvGQ/Mb9p4LLLsjAUjBhgXqEK5Im2kr70bJWjJBx4rJounCBgmlGXkeCaOaHDBGV7wA7lFc7IBBOAShYbkohrPkOOqCfBATlzNrWCzmN8N8FTxN3Y2t7gpvQI+lv7547oz7VTKsQYLlqKrHvykgegZWp55pfSSfV/mty1VhstDZmbwBzzIWv5po/QD3j8VLvsDBLrdYDqdI//wAgfik1+5Jbky7g0cmp+dKogE9APePxU9qDY2W5GRN3WNOAMqsr7gb7MxYhkmPNzsD1D+6crU0mLoaMTO3dyfWeP4rbV1yQPmRu8wt7wpmLT3yai+EDDGu4n0KoXgaAc44qGrB8V4WQRhjAAAFkQhWAu2gl6Og4Z4vw0e3+y+NnYAyoJcdZ5zlaW0su9OyIHg3JK29nrTHVhAThzOz0Kn9wGx4BKrmrivadE6GTDTzyBlNVhnrQzjErA71qz+QNenqdawQ1r91x+a7gfzWHWtQEMfRxOy9w8Fgu6KAC+s4g9xPBJrAlbK5s29vjgcqjbQPgkk5PNDQXEADJPJeJxs/S6R/lMg6o+T8VVK2Bjo9MVoAXDru4lb6OAHcAtOpdFi5LE3G6wDisvQG475JHoU7UkFXamGXJDW2Wn2ZVEpjWgYtU3/6XBY80FODi+5lwZHiyxmuqafg6+ODe/sXKNs522No7Tm43Wu3B7F0vytrdH8sPACHf+zK4/YkM08krjkvcXFcR6W0zWXJkfbl/3g+w+0/cVPS6bBF/1ff/AGql/tmNCELtT4yC2qliyInUo53shmcN9oOAcLVQgKevpNMRDei3ye0krHLold3yN5p9BXxoupNewQTHDxyPenAIIyCsqSYEzKOo1SDT1CWPHLDi37l7Z1DanydzJdStvixxAnyMeKcJfr0gjov73cB7VVwXUCTRozLqDTz3cu9vZ9qqwMDA5JFsxEP3kp7wAnqmPQCnaSbdrCIHi4rzZqHdrOlPN7vsH6K0topd+6Ix8wfaU80+IQ044+5oHx+3KLmwbBOASlNG86TU5onPJZ83Po4FMbsnRVnv7gSpKrMYrbJs8nZKSfNAsiQBk8knsXnWb8daAkMDsuI7cLHrWo/u+ghd1iOse4L42ah3pZJiOWGj7z+ClvnQH4GAAlW0ku5T6MHi9wH4/BNVNbQziW2I2nIZz9ZSXQG7s1EBXdKRxc7HsCcHgMrW0yLoaUTORDRn181skZGCpXQCyXWYI5nxuDuqccBlfPnyt3Se7+aynR6pcXFuSTnt+K88z1P4B4n4qPvAwya3XLCGh+f6fzSWsw2LrGn57+P4p1c0upDA+Td+SM9vxWhs/F0lwvxwaPv/AEVVp3TBSsGGALQ1yw6Cr1HEOJGMJgkG00uZWRA8uKvJ0gxzSlE1Zkg5EZXxqFtlSEvcRvdg70r0O8yGnIyQ46M5HqP5/elt2zJcs7x5E4aFXi5Af6M6SaEzyEkvJIHcEwPAFYaUYirMYByAC+NSnbBVe4njjgrdgTOpydLekcDkA4HsX0ytbhibaYCO3hzAWbRqflVjpJOLGnOO8qm3G7u7jgqKN8wIamtObhs7M+kJvUuQWR+7eCe5YLWlVZiXbm649rTheUdLhqz9K1zycdpVlYGCRbTQtBZKBg5wU9SDaKYSzR14+LgeP4JLoBdQrOtWGxtHD5x9CrYI2wxNjaAABhamkUxVrjPyzxJW8SBxKRVIGvqIndWc2vjfPDicJbo1KzXtudJu7pbxwc9qc7zR/ZAcO9TQPVP7TsxPFJjmCP14qgSnaZmabX/wvH2g/BRLoBrqGp42EgYH9eRoj9n6CiFu2b3S6VWpBmOhc5xdnnlaS1e36RaaM18ZN+XyN7vu6vcsmKV8oQhHwuf+bBCEL3miBCEIAHA5CZ0dXmhAbLl7R29qWIUp0Coh1epIOMgaf+wwl20FuOfo44ntcOZwUoQpcmwUWjWKsFJjXTxtceJyVuO1Cpg4sReKkUIpMG30jJ9T6SRwDC/JJ5YVE2/UDQPKIveUkhFKgUGt3YX0yyKVjy7h1Sp9CFDdgE+0axVrUhvzMDjkkZ4/rgEhQidAfahrLNwsr9Yn52OCT1t2S3H0rgGl4LiVhQjdgrW36gH+oi8V75wqfzEXipFCtxsFb5xqfTx+KBqNP+Yj8VJITjYKDW7sL6RZFKx5ccYaVh0CWtBE50kzGuceTikqFHFzsFcb9Mf8iL3lN6pMJ7j3tOW8gVqoRysAtjT9zyyMyODWg5JK10KoKmTVKkTMCVriBybxSLUr0lt/HgwcgtNCs5Ngy1rEtd+9G4jvHYU8qa1E4ATAsPf2KeQik0CwZdrPA3ZWcf8AsvZLldg60rB/6CjkKeMD+/rLA0sr9Z3f2LS0p0PTmzambvZ4Anj60tQo4gVvnGmP9+PxSvWtREgEVeTgeZaUmQjkwZfKJ/ppPeKz6dbfHcjfLK4sB47zjhaaFUkrPONTH+fH4rS1m3WsUnMZM0u5gZSBCtxEAhCFUAhCEB//2Q==" '
        'style="width:100%;max-width:180px;display:block;margin:0 auto;" alt="IMEMSA"/>'
        '<div style="height:2px;background:#C41E2E;border-radius:1px;margin-top:10px;"></div>'
        '</div>',
        unsafe_allow_html=True,
    )


def logo_login() -> None:
    st.markdown(
        '<div style="text-align:center;margin-bottom:24px;">'
        '<div style="display:inline-block;background:#0D2B6E;border-radius:10px;padding:14px 24px 10px;">'
        + _logo_img_tag(width=180) +
        '<div style="font-family:Source Sans 3,Arial,sans-serif;font-size:.7rem;'
        'font-weight:600;color:#8EA8D8;letter-spacing:2.5px;margin-top:5px;">MOTORES YAMAHA</div>'
        '<div style="height:2px;background:#C41E2E;border-radius:1px;margin-top:5px;"></div>'
        '</div></div>',
        unsafe_allow_html=True,
    )


# ══════════════════════════════════════════════════════════
#  HELPERS UI
# ══════════════════════════════════════════════════════════

def _sem_dot(color: str) -> str:
    return f'<span class="semaphore-dot sem-{color}"></span>'

def _progress_bar(pct: int) -> str:
    return (f'<div class="progress-wrap">'
            f'<div class="progress-fill" style="width:{pct}%;"></div>'
            f'</div><span style="font-size:.75rem;color:#8592A3;">{pct}% completado</span>')

def _badge(status: str) -> str:
    cls = {"active": "badge-active", "completed": "badge-completed",
           "cancelled": "badge-cancelled"}.get(status, "badge-active")
    return f'<span class="badge {cls}">{ORDER_STATUS_LABELS.get(status, status)}</span>'

def _avatar(initials: str, color: str) -> str:
    return f'<span class="avatar" style="background:{color};">{initials}</span>'

def _act_row_class(status: str) -> str:
    return {"completed": "act-row-completed", "in_progress": "act-row-inprogress",
            "waiting_closure": "act-row-waiting", "pending": "act-row-pending",
            "blocked": "act-row-blocked"}.get(status, "act-row-pending")


# ══════════════════════════════════════════════════════════
#  HELPERS — MEJORAS v2.1
# ══════════════════════════════════════════════════════════

def _working_days_remaining(due_date_str: str) -> int | None:
    """Días hábiles (lun-vie) entre hoy y la fecha límite. Negativo = vencida."""
    if not due_date_str:
        return None
    try:
        due   = datetime.strptime(due_date_str, "%Y-%m-%d").date()
        today = date.today()
        if due == today:
            return 0
        step, days, current = (1, 0, today) if due > today else (-1, 0, today)
        while current != due:
            current += timedelta(days=step)
            if current.weekday() < 5:
                days += step
        return days
    except Exception:
        return None


def _days_badge(days: int | None) -> str:
    if days is None:
        return ""
    if days > 2:
        color, bg, txt = "#065F46", "#D1FAE5", f"⏱ {days}d restantes"
    elif days >= 0:
        color, bg, txt = "#92400E", "#FEF3C7", f"⚠️ {days}d restantes"
    else:
        color, bg, txt = "#991B1B", "#FEE2E2", f"🔴 {abs(days)}d vencida"
    return (f' <span style="font-size:.68rem;font-weight:700;padding:2px 8px;'
            f'border-radius:12px;background:{bg};color:{color};">{txt}</span>')


def _log_history(act: dict, username: str, action: str, detail: str = "") -> None:
    """Agrega entrada al historial de una actividad (se guarda con _app_save)."""
    if "history" not in act:
        act["history"] = []
    act["history"].append({
        "ts":     datetime.now().strftime("%d/%m/%Y %H:%M"),
        "user":   USERS.get(username, {}).get("name", username),
        "action": action,
        "detail": detail,
    })


# ══════════════════════════════════════════════════════════
#  GOOGLE DRIVE — GALERÍA DE EVIDENCIAS
# ══════════════════════════════════════════════════════════

def _get_drive_client():
    """Regresa cliente Google Drive autenticado con el mismo service account."""
    import streamlit as st
    from google.oauth2.service_account import Credentials
    import googleapiclient.discovery as discovery
    scopes = ["https://www.googleapis.com/auth/drive"]
    creds  = Credentials.from_service_account_info(
        dict(st.secrets["gcp_service_account"]), scopes=scopes
    )
    return discovery.build("drive", "v3", credentials=creds, cache_discovery=False)


def _get_or_create_drive_folder(drive, folder_name: str = "IMEMSA_Evidencias") -> str:
    """Obtiene o crea la carpeta de evidencias en Google Drive. Retorna el folder_id."""
    # Buscar si ya existe
    q = f"name='{folder_name}' and mimeType='application/vnd.google-apps.folder' and trashed=false"
    results = drive.files().list(q=q, fields="files(id,name)").execute()
    files   = results.get("files", [])
    if files:
        return files[0]["id"]
    # Crear carpeta nueva
    meta = {"name": folder_name, "mimeType": "application/vnd.google-apps.folder"}
    folder = drive.files().create(body=meta, fields="id").execute()
    # Hacer la carpeta pública para lectura (todos con el link pueden ver)
    drive.permissions().create(
        fileId=folder["id"],
        body={"type": "anyone", "role": "reader"},
    ).execute()
    return folder["id"]


def _upload_evidence_to_drive(
    file_bytes: bytes,
    file_name:  str,
    order_number: str,
    act_id: int,
) -> tuple[str, str]:
    """
    Sube la evidencia a Google Drive.
    Retorna (file_id, view_url, thumb_url) o ("", "", "") si falla.
    """
    try:
        import io as _io
        from googleapiclient.discovery import build       # noqa
        from googleapiclient.http import MediaIoBaseUpload

        drive     = _get_drive_client()
        folder_id = _get_or_create_drive_folder(drive)

        # Detectar mime type
        ext = file_name.lower().split(".")[-1]
        mime_map = {
            "pdf":  "application/pdf",
            "png":  "image/png",
            "jpg":  "image/jpeg",
            "jpeg": "image/jpeg",
            "xlsx": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            "docx": "application/vnd.openxmlformats-officedocument.wordprocessingml.document",
        }
        mime_type = mime_map.get(ext, "application/octet-stream")

        # Nombre del archivo con contexto
        safe_order = order_number.replace("/", "-").replace(" ", "_")
        drive_name = f"{safe_order}_Act{act_id:02d}_{file_name}"

        # Subir archivo
        media = MediaIoBaseUpload(_io.BytesIO(file_bytes), mimetype=mime_type)
        meta  = {"name": drive_name, "parents": [folder_id]}
        uploaded = drive.files().create(
            body=meta, media_body=media, fields="id"
        ).execute()
        file_id = uploaded["id"]

        # Permisos públicos de lectura
        drive.permissions().create(
            fileId=file_id,
            body={"type": "anyone", "role": "reader"},
        ).execute()

        view_url = f"https://drive.google.com/file/d/{file_id}/view"
        thumb_url = f"https://drive.google.com/thumbnail?id={file_id}&sz=w400"
        print(f"[DRIVE UPLOAD OK] {drive_name} → {view_url}")
        return file_id, view_url, thumb_url

    except ImportError as e:
        import streamlit as _st
        _st.error(f"❌ Drive — Falta librería: {e}\nAgrega al requirements.txt y redespliega.")
        return "", "", ""
    except Exception as e:
        import streamlit as _st
        _st.error(f"❌ Drive error ({type(e).__name__}): {e}")
        return "", "", ""


def _render_evidence_gallery(act: dict) -> None:
    """Muestra la galería de evidencias de una actividad — visible para todos los usuarios."""
    evidences = act.get("evidences", [])

    # Compatibilidad: evidencia antigua con URL de Drive
    if not evidences and act.get("evidence_name") and act.get("evidence_url"):
        evidences = [{
            "name":      act["evidence_name"],
            "url":       act["evidence_url"],
            "thumb_url": act.get("evidence_thumb", ""),
            "user":      "—",
            "ts":        act.get("completion_date", "—"),
        }]

    # Compatibilidad: evidencia antigua SIN URL (solo nombre de archivo)
    if not evidences and act.get("evidence_name"):
        evidences = [{
            "name":      act["evidence_name"],
            "url":       "",
            "thumb_url": "",
            "user":      "—",
            "ts":        act.get("completion_date", "—"),
            "sin_url":   True,  # bandera para mostrar aviso
        }]

    if not evidences:
        return

    st.markdown(
        f'<div style="margin-top:8px;padding:10px 14px;background:#F0F7FF;'
        f'border-radius:8px;border-left:3px solid #0D2B6E;">'
        f'<span style="font-size:.78rem;font-weight:700;color:#0D2B6E;">'
        f'📸 Evidencias ({len(evidences)})</span></div>',
        unsafe_allow_html=True,
    )
    cols = st.columns(min(len(evidences), 3))
    for i, ev in enumerate(evidences):
        with cols[i % 3]:
            thumb = ev.get("thumb_url", "")
            url   = ev.get("url", "")
            name  = ev.get("name", "archivo")
            user  = ev.get("user", "—")
            ts    = ev.get("ts", "—")
            ext   = name.lower().split(".")[-1] if "." in name else ""

            # ── Sin URL: upload falló o evidencia antigua ─────────────
            if ev.get("sin_url"):
                st.markdown(
                    f'<div style="padding:10px 14px;background:#FFF8F0;'
                    f'border:1.5px solid #F59E0B;border-radius:8px;text-align:center;">'
                    f'<div style="font-size:1.2rem;">⚠️</div>'
                    f'<div style="font-size:.72rem;color:#92400E;font-weight:600;margin-top:4px;">{name}</div>'
                    f'<div style="font-size:.68rem;color:#6B7280;">Subida a Drive pendiente<br>👤 {user} · {ts}</div>'
                    f'</div>',
                    unsafe_allow_html=True,
                )
            elif thumb and ext in ("png", "jpg", "jpeg"):
                # Mostrar miniatura con link
                st.markdown(
                    f'<a href="{url}" target="_blank">'
                    f'<img src="{thumb}" style="width:100%;border-radius:8px;'
                    f'border:2px solid #D1D9E8;margin-bottom:4px;" alt="{name}"/></a>'
                    f'<div style="font-size:.70rem;color:#6B7280;">'
                    f'📎 {name}<br>👤 {user} · {ts}</div>',
                    unsafe_allow_html=True,
                )
            else:
                # PDF / Excel — mostrar link
                st.markdown(
                    f'<a href="{url}" target="_blank" style="text-decoration:none;">'
                    f'<div style="padding:10px 14px;background:#fff;border:1.5px solid #D1D9E8;'
                    f'border-radius:8px;text-align:center;">'
                    f'<div style="font-size:1.4rem;">📄</div>'
                    f'<div style="font-size:.72rem;color:#374151;font-weight:600;margin-top:4px;">{name}</div>'
                    f'<div style="font-size:.68rem;color:#6B7280;">👤 {user} · {ts}</div>'
                    f'</div></a>',
                    unsafe_allow_html=True,
                )


def _export_order_excel(order: dict) -> bytes:
    """Genera Excel con el detalle de actividades del pedido."""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
        wb  = openpyxl.Workbook()
        ws  = wb.active
        ws.title = "Actividades"
        navy, red_c, white = "0D2B6E", "C41E2E", "FFFFFF"
        # Fila 1 — título
        ws.merge_cells("A1:H1")
        c = ws["A1"]
        c.value     = f"IMEMSA  ·  {order['order_number']}  —  {order['motor_model']}"
        c.font      = Font(bold=True, color=white, size=13)
        c.fill      = PatternFill("solid", fgColor=navy)
        c.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[1].height = 28
        # Fila 2 — resumen
        ws.merge_cells("A2:H2")
        c = ws["A2"]
        c.value     = (f"Cantidad: {order['quantity']} u.  ·  Monto/OC: {order['supplier']}"
                       f"  ·  Creado: {order['created_at']}  ·  Avance: {order.get('progress',0)}%")
        c.font      = Font(size=10, color="555555")
        c.alignment = Alignment(horizontal="center")
        ws.row_dimensions[2].height = 16
        # Fila 3 — encabezados columnas
        headers = ["#", "Fase", "Actividad", "Responsable", "Estado",
                   "Inicio", "Límite", "Cierre"]
        for col, h in enumerate(headers, 1):
            c = ws.cell(row=3, column=col, value=h)
            c.font      = Font(bold=True, color=white, size=10)
            c.fill      = PatternFill("solid", fgColor=red_c)
            c.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[3].height = 20
        status_map = {"completed": "Completada", "in_progress": "En curso",
                      "waiting_closure": "Esp. cierre", "pending": "Pendiente",
                      "blocked": "Bloqueada"}
        fills = {"completed": "F0FDF4", "in_progress": "EFF6FF",
                 "waiting_closure": "FFFBEB", "blocked": "FEF2F2"}
        for i, act in enumerate(order.get("activities", []), 4):
            resp = USERS.get(act.get("responsible_key", ""), {})
            st_  = act.get("status", "pending")
            row  = [act.get("id",""), act.get("phase",""), act.get("name",""),
                    resp.get("name",""), status_map.get(st_, st_),
                    act.get("start_date",""), act.get("due_date",""),
                    act.get("completion_date","")]
            bg = fills.get(st_, "FFFFFF")
            for col, val in enumerate(row, 1):
                c = ws.cell(row=i, column=col, value=val)
                c.fill      = PatternFill("solid", fgColor=bg)
                c.alignment = Alignment(vertical="center", wrap_text=True)
            ws.row_dimensions[i].height = 18
        for col, w in zip("ABCDEFGH", [5, 24, 38, 22, 14, 13, 13, 13]):
            ws.column_dimensions[col].width = w
        buf = io.BytesIO()
        wb.save(buf)
        return buf.getvalue()
    except Exception:
        return b""


# ══════════════════════════════════════════════════════════
#  ALERTA AUTOMÁTICA DIARIA (8:00 am)
# ══════════════════════════════════════════════════════════

def _check_daily_alerts() -> None:
    """Se ejecuta una vez por día al detectar la primera sesión activa a partir de las 8am."""
    now       = datetime.now()
    today_str = now.strftime("%Y-%m-%d")
    last      = st.session_state.get("_last_alert_check", "")
    if now.hour >= 8 and last != today_str:
        st.session_state["_last_alert_check"] = today_str
        data = _app_data()
        reds = get_red_activities(data)
        if reds:
            sent = 0
            for act in reds:
                ok = send_overdue_alert(
                    to_email=act["_responsible_email"],
                    to_name=act["_responsible_name"],
                    order_number=act["_order_number"],
                    activity_name=act["name"],
                    due_date=act.get("due_date", "—"),
                )
                if ok:
                    sent += 1
            if sent > 0:
                st.toast(
                    f"🔔 Alerta automática 8am: {sent} actividad(es) vencida(s) notificadas.",
                    icon="📧",
                )


# ══════════════════════════════════════════════════════════
#  PÁGINA 1 — LOGIN
# ══════════════════════════════════════════════════════════

def page_login() -> None:
    st.markdown("""
    <style>
    @import url('https://fonts.googleapis.com/css2?family=Bebas+Neue&family=Rajdhani:wght@300;400;600;700&display=swap');

    [data-testid="stAppViewContainer"] {
        background: radial-gradient(ellipse 90% 70% at 50% -5%,
            #1a3a8f 0%, #0a1d5e 28%, #020b22 65%) !important;
        min-height: 100vh;
    }
    #MainMenu, footer, header { visibility: hidden; }
    .main .block-container { padding-top: 0 !important; }

    [data-testid="stAppViewContainer"]::before {
        content: ''; position: fixed; inset: 0; z-index: 0; pointer-events: none;
        background: repeating-linear-gradient(
            0deg, transparent, transparent 3px,
            rgba(0,0,0,0.07) 3px, rgba(0,0,0,0.07) 4px);
    }

    [data-testid="stAppViewContainer"] [data-testid="stVerticalBlock"],
    [data-testid="stAppViewContainer"] [data-testid="column"] > div:first-child,
    div[data-testid="stVerticalBlockBorderWrapper"] {
        background: transparent !important; border: none !important; box-shadow: none !important;
    }

    /* Labels en blanco */
    [data-testid="stTextInput"] label,
    [data-testid="stTextInput"] label p,
    [data-testid="stTextInput"] label span,
    [data-testid="stWidgetLabel"] p,
    [data-testid="stWidgetLabel"] span {
        color: #ffffff !important;
        font-family: 'Rajdhani', sans-serif !important;
        font-size: 13px !important; letter-spacing: 3px !important;
        text-transform: uppercase !important;
    }

    [data-testid="stTextInput"] input {
        background: rgba(255,255,255,0.93) !important;
        border: none !important; border-radius: 7px !important;
        color: #1a2a5e !important;
        font-family: 'Rajdhani', sans-serif !important; font-size: 15px !important;
        transition: box-shadow 0.25s !important;
    }
    [data-testid="stTextInput"] input:focus { box-shadow: 0 0 0 2px #C41E2E !important; }
    [data-testid="stTextInput"] input::placeholder { color: #8899bb !important; }

    [data-testid="stForm"] {
        background: rgba(13,43,110,0.45) !important;
        border: 1px solid rgba(100,150,255,0.18) !important;
        border-radius: 12px !important;
        padding: 10px 14px !important;
        backdrop-filter: blur(10px) !important;
    }

    [data-testid="stFormSubmitButton"] button {
        width: 100% !important;
        background: linear-gradient(135deg, #d42030 0%, #a8151f 100%) !important;
        border: none !important; border-radius: 8px !important; color: #ffffff !important;
        font-family: 'Bebas Neue', sans-serif !important;
        font-size: 20px !important; letter-spacing: 5px !important; padding: 13px !important;
        box-shadow: 0 4px 20px rgba(196,30,46,0.55) !important; transition: all 0.25s !important;
    }
    [data-testid="stFormSubmitButton"] button:hover {
        transform: translateY(-2px) !important;
        box-shadow: 0 6px 28px rgba(196,30,46,0.7) !important;
    }

    [data-testid="stAlert"] {
        background: rgba(196,30,46,0.12) !important;
        border: 1px solid rgba(196,30,46,0.35) !important;
        border-radius: 8px !important; color: #f07080 !important;
    }

    .corner-tl, .corner-tr, .corner-bl, .corner-br {
        position: fixed; width: 40px; height: 40px; z-index: 100; pointer-events: none;
    }
    .corner-tl { top:16px; left:16px;  border-top:2px solid #C41E2E; border-left:2px solid #C41E2E; }
    .corner-tr { top:16px; right:16px; border-top:2px solid #C41E2E; border-right:2px solid #C41E2E; }
    .corner-bl { bottom:16px; left:16px;  border-bottom:2px solid #C41E2E; border-left:2px solid #C41E2E; }
    .corner-br { bottom:16px; right:16px; border-bottom:2px solid #C41E2E; border-right:2px solid #C41E2E; }

    @keyframes rise-bubble {
        0%   { transform: translateY(0); opacity: 0.65; }
        80%  { opacity: 0.15; }
        100% { transform: translateY(-105vh); opacity: 0; }
    }
    .bbl {
        position: fixed; border-radius: 50%;
        background: rgba(180,210,255,0.20); border: 1px solid rgba(180,210,255,0.32);
        animation: rise-bubble linear infinite; pointer-events: none; z-index: 1; bottom: -20px;
    }
    </style>

    <div class="corner-tl"></div><div class="corner-tr"></div>
    <div class="corner-bl"></div><div class="corner-br"></div>

    <div class="bbl" style="width:5px;height:5px;left:4%;animation-duration:9s;animation-delay:-2s;"></div>
    <div class="bbl" style="width:8px;height:8px;left:10%;animation-duration:14s;animation-delay:-7s;"></div>
    <div class="bbl" style="width:4px;height:4px;left:17%;animation-duration:11s;animation-delay:-1s;"></div>
    <div class="bbl" style="width:10px;height:10px;left:24%;animation-duration:16s;animation-delay:-11s;"></div>
    <div class="bbl" style="width:5px;height:5px;left:31%;animation-duration:8s;animation-delay:-4s;"></div>
    <div class="bbl" style="width:7px;height:7px;left:38%;animation-duration:13s;animation-delay:-9s;"></div>
    <div class="bbl" style="width:3px;height:3px;left:45%;animation-duration:10s;animation-delay:-0s;"></div>
    <div class="bbl" style="width:9px;height:9px;left:52%;animation-duration:17s;animation-delay:-6s;"></div>
    <div class="bbl" style="width:4px;height:4px;left:59%;animation-duration:12s;animation-delay:-13s;"></div>
    <div class="bbl" style="width:6px;height:6px;left:66%;animation-duration:9s;animation-delay:-3s;"></div>
    <div class="bbl" style="width:11px;height:11px;left:72%;animation-duration:15s;animation-delay:-8s;"></div>
    <div class="bbl" style="width:4px;height:4px;left:79%;animation-duration:11s;animation-delay:-5s;"></div>
    <div class="bbl" style="width:7px;height:7px;left:85%;animation-duration:13s;animation-delay:-10s;"></div>
    <div class="bbl" style="width:5px;height:5px;left:91%;animation-duration:8s;animation-delay:-15s;"></div>
    <div class="bbl" style="width:9px;height:9px;left:97%;animation-duration:16s;animation-delay:-2s;"></div>
    <div class="bbl" style="width:3px;height:3px;left:13%;animation-duration:10s;animation-delay:-18s;"></div>
    <div class="bbl" style="width:6px;height:6px;left:55%;animation-duration:12s;animation-delay:-16s;"></div>
    <div class="bbl" style="width:8px;height:8px;left:42%;animation-duration:7s;animation-delay:-12s;"></div>

    <div style="position:fixed;bottom:18px;right:22px;z-index:100;
        font-size:9px;letter-spacing:3px;color:rgba(200,216,240,0.3);
        text-transform:uppercase;font-family:'Rajdhani',sans-serif;">
        Sistema v2.0 · 2026</div>
    """, unsafe_allow_html=True)

    st.markdown("<div style='height:5vh'></div>", unsafe_allow_html=True)

    col_l, col_c, col_r = st.columns([1, 1.1, 1])
    with col_c:

        st.markdown("""
        <div style="background:rgba(13,43,110,0.60);border:1px solid rgba(100,150,255,0.20);
            border-radius:12px;padding:22px 24px 18px;text-align:center;
            backdrop-filter:blur(8px);margin-bottom:20px;">
            <div style="font-family:'Bebas Neue',sans-serif;font-size:46px;letter-spacing:6px;
                color:#ffffff;text-shadow:0 0 30px rgba(100,150,255,0.5);line-height:1;">IMEMSA</div>
            <div style="width:200px;height:2px;margin:10px auto;
                background:linear-gradient(to right,transparent,#C41E2E 30%,#C41E2E 70%,transparent);
                box-shadow:0 0 10px rgba(196,30,46,0.55);"></div>
            <div style="font-size:11px;letter-spacing:6px;color:#c8d8f0;opacity:0.7;
                text-transform:uppercase;font-family:'Rajdhani',sans-serif;">Motores Yamaha</div>
        </div>
        """, unsafe_allow_html=True)

        st.markdown("""
        <p style="text-align:center;margin-bottom:18px;font-size:14px;letter-spacing:1px;
            color:#c8d8f0;line-height:1.6;font-family:'Rajdhani',sans-serif;">
            Sistema de Seguimiento ·
            <span style="color:#ffffff;font-weight:600;">Proceso Transversal</span><br>
            de Compra de Motores
        </p>
        """, unsafe_allow_html=True)

        with st.form("login_form"):
            username  = st.text_input("👤  Usuario",    placeholder="Ej. dgonzalez")
            password  = st.text_input("🔒  Contraseña", type="password")
            submitted = st.form_submit_button("Ingresar al Sistema",
                                              use_container_width=True, type="primary")

        if submitted:
            ok, user_info = verify_login(username, password)
            if ok:
                st.session_state["user"]              = user_info
                st.session_state["page"]              = "dashboard"
                st.session_state["selected_order_id"] = None
                st.rerun()
            else:
                st.error("Usuario o contraseña incorrectos.")

        st.markdown(
            '<p style="text-align:center;color:rgba(255,255,255,.30);font-size:.72rem;'
            'margin-top:20px;font-family:\'Rajdhani\',sans-serif;">'
            '© 2026 IMEMSA — Uso interno. Acceso restringido.</p>',
            unsafe_allow_html=True,
        )


# ══════════════════════════════════════════════════════════
#  SIDEBAR NAVEGACIÓN
# ══════════════════════════════════════════════════════════

def render_sidebar() -> str:
    user = st.session_state["user"]
    page = st.session_state.get("page", "dashboard")

    with st.sidebar:
        logo_sidebar()
        st.markdown('<div class="sidebar-divider"></div>', unsafe_allow_html=True)

        av = _avatar(user["avatar"], user["color"])
        st.markdown(
            f'<div style="padding:4px 16px 12px;">'
            f'{av}<span style="font-size:.9rem;font-weight:700;">{user["name"]}</span>'
            f'<br><span style="font-size:.75rem;padding-left:46px;opacity:.7;">{user["role"]}</span>'
            f'</div>',
            unsafe_allow_html=True,
        )
        st.markdown('<div class="sidebar-divider"></div>', unsafe_allow_html=True)

        st.markdown(
            '<p style="font-size:.7rem;font-weight:700;text-transform:uppercase;'
            'letter-spacing:1px;padding-left:4px;opacity:.55;margin-bottom:6px;">Navegación</p>',
            unsafe_allow_html=True,
        )

        nav_items = [
            ("📊", "Tablero de Pedidos", "dashboard"),
            ("📋", "Detalle de Actividades", "activities"),
        ]
        if user.get("can_create_orders"):
            nav_items.append(("➕", "Nuevo Pedido", "new_order"))

        for icon, label, key in nav_items:
            active_style = "nav-active" if page == key else ""
            st.markdown(f'<div class="{active_style}">', unsafe_allow_html=True)
            if st.button(f"{icon}  {label}", key=f"nav_{key}", use_container_width=True):
                st.session_state["page"] = key
                st.session_state["selected_order_id"] = None
                st.rerun()
            st.markdown("</div>", unsafe_allow_html=True)

        st.markdown('<div class="sidebar-divider"></div>', unsafe_allow_html=True)

        data = _app_data()
        sem  = get_semaphore_summary(data)
        st.markdown(
            '<p style="font-size:.7rem;font-weight:700;text-transform:uppercase;'
            'letter-spacing:1px;padding-left:4px;opacity:.55;margin-bottom:8px;">Estado General</p>',
            unsafe_allow_html=True,
        )
        st.markdown(
            f'<div style="background:rgba(255,255,255,.07);border-radius:10px;padding:12px 14px;">'
            f'{_sem_dot("green")}<span style="font-size:.85rem;">{sem["green"]} En tiempo</span><br>'
            f'<div style="margin-top:5px;">{_sem_dot("yellow")}<span style="font-size:.85rem;">{sem["yellow"]} En riesgo</span></div><br>'
            f'<div style="margin-top:-8px;">{_sem_dot("red")}<span style="font-size:.85rem;">{sem["red"]} Vencidas</span></div>'
            f'</div>',
            unsafe_allow_html=True,
        )

        if sem["red"] > 0:
            st.markdown('<div style="margin-top:10px;"></div>', unsafe_allow_html=True)
            if st.button("🚨  Enviar alertas vencidas", use_container_width=True):
                _send_overdue_alerts(_app_data())

        st.markdown('<div class="sidebar-divider"></div>', unsafe_allow_html=True)

        if st.button("🚪  Cerrar sesión", use_container_width=True):
            _app_invalidate()
            for k in list(st.session_state.keys()):
                del st.session_state[k]
            st.rerun()

    return page


def _send_overdue_alerts(data: dict) -> None:
    reds = get_red_activities(data)
    sent = 0
    for act in reds:
        ok = send_overdue_alert(
            to_email=act["_responsible_email"], to_name=act["_responsible_name"],
            order_number=act["_order_number"], activity_name=act["name"],
            due_date=act.get("due_date", "—"),
        )
        if ok:
            sent += 1
    if sent:
        st.toast(f"✅ {sent} alerta(s) enviadas.", icon="📧")
    else:
        st.toast("No se pudo enviar correos (verifica SMTP).", icon="⚠️")


# ══════════════════════════════════════════════════════════
#  PÁGINA 2 — TABLERO DE PEDIDOS
# ══════════════════════════════════════════════════════════

def page_dashboard() -> None:
    data         = _app_data()
    user         = st.session_state["user"]
    orders       = get_orders_for_user(data, user["username"])
    active       = [o for o in orders if o["status"] == "active"]
    completed    = [o for o in orders if o["status"] == "completed"]
    sem          = get_semaphore_summary(data)
    year         = datetime.today().year
    annual_count = len(orders)  # todos los pedidos activos + completados (no cancelados)

    st.markdown(
        f'<div style="display:flex;align-items:center;justify-content:space-between;margin-bottom:6px;">'
        f'<div class="section-header" style="margin-bottom:0;">📊 Tablero de Pedidos</div>'
        f'<span style="font-size:.82rem;color:#8592A3;">Actualizado: {datetime.today().strftime("%d/%m/%Y %H:%M")}</span>'
        f'</div>',
        unsafe_allow_html=True,
    )

    c1, c2, c3, c4, c5 = st.columns(5)
    _kpi(c1, str(len(active)),    "Pedidos Activos",   "#0D2B6E")
    _kpi(c2, str(len(completed)), "Pedidos Completados",        "#059669")
    _kpi(c3, str(annual_count),   f"Pedidos {year}",   "#7C3AED")
    _kpi(c4, str(sem["red"]),     "Actividades Vencidas",  "#C41E2E")
    _kpi(c5, str(sem["yellow"]),  "Actividades en Riesgo",          "#D97706")

    st.markdown("<br>", unsafe_allow_html=True)

    if sem["red"] > 0:
        st.markdown(
            f'<div style="background:#FEF2F2;border:1.5px solid #FECACA;border-radius:10px;'
            f'padding:12px 18px;margin-bottom:16px;display:flex;align-items:center;gap:10px;">'
            f'{_sem_dot("red")}'
            f'<span style="color:#991B1B;font-weight:700;font-size:.88rem;">'
            f'{sem["red"]} actividad(es) vencida(s). Usa "Enviar alertas" en el menú lateral.</span>'
            f'</div>',
            unsafe_allow_html=True,
        )

    my_tasks = get_my_pending_activities(data, user["username"])
    if my_tasks:
        st.markdown('<div class="section-header" style="font-size:1.1rem;">🔔 Mis tareas activas</div>',
                    unsafe_allow_html=True)
        for t in my_tasks:
            sem_color = get_semaphore(t)
            st.markdown(
                f'<div class="act-row act-row-inprogress" style="margin-bottom:6px;">'
                f'{_sem_dot(sem_color)}<span class="act-name">{t["name"]}</span>'
                f'<span class="phase-chip">{t["phase"]}</span>'
                f'<span class="act-meta"> · {t["_order_number"]} · Vence: {t.get("due_date","—")}</span>'
                f'</div>',
                unsafe_allow_html=True,
            )
        st.markdown("<br>", unsafe_allow_html=True)

    # ── FILTROS ──────────────────────────────────────────────
    st.markdown('<div class="section-header" style="font-size:1.1rem;">🔍 Filtros</div>',
                unsafe_allow_html=True)
    resp_names = sorted({
        USERS[a.get("responsible_key","")]["name"]
        for o in orders
        for a in o.get("activities", [])
        if a.get("responsible_key","") in USERS
    })
    col_f1, col_f2, _ = st.columns([1, 1, 2])
    with col_f1:
        filter_resp = st.selectbox("Responsable", ["Todos"] + resp_names, key="f_resp")
    with col_f2:
        filter_sem  = st.selectbox("Semáforo activo", ["Todos", "🟢 En tiempo", "🟡 En riesgo", "🔴 Vencida"], key="f_sem")
    sem_map = {"🟢 En tiempo": "green", "🟡 En riesgo": "yellow", "🔴 Vencida": "red"}

    def _pass_filters(order: dict) -> bool:
        acts = order.get("activities", [])
        active_act = next((a for a in acts if a["status"] in ("in_progress","waiting_closure")), None)
        if filter_resp != "Todos":
            resp_key = next((k for k,v in USERS.items() if v["name"] == filter_resp), None)
            if not any(a.get("responsible_key") == resp_key for a in acts
                       if a["status"] in ("in_progress","waiting_closure")):
                return False
        if filter_sem != "Todos" and active_act:
            if get_semaphore(active_act) != sem_map.get(filter_sem, ""):
                return False
        return True

    active_f    = [o for o in active    if _pass_filters(o)]
    completed_f = [o for o in completed if _pass_filters(o)]

    tabs = st.tabs([f"🔄  Activos  ({len(active_f)})", f"✅  Completados  ({len(completed_f)})"])
    with tabs[0]:
        if not active_f:
            st.info("No hay pedidos activos con los filtros seleccionados.")
        else:
            for order in sorted(active_f, key=lambda o: o["id"], reverse=True):
                _render_order_card(order)
    with tabs[1]:
        if not completed_f:
            st.info("No hay pedidos completados con los filtros seleccionados.")
        else:
            for order in sorted(completed_f, key=lambda o: o["id"], reverse=True):
                _render_order_card(order)


def _kpi(col, value: str, label: str, color: str) -> None:
    with col:
        st.markdown(
            f'<div class="metric-card" style="--card-accent:{color};">'
            f'<div class="mc-value">{value}</div>'
            f'<div class="mc-label">{label}</div>'
            f'</div>',
            unsafe_allow_html=True,
        )


def _render_order_card(order: dict) -> None:
    acts            = order.get("activities", [])
    active_act      = next((a for a in acts if a["status"] in ("in_progress", "waiting_closure")), None)
    responsible_key = active_act["responsible_key"] if active_act else None
    responsible     = USERS.get(responsible_key, {})
    prog            = order.get("progress", 0)
    sem_color       = get_semaphore(active_act) if active_act else "gray"

    with st.container():
        st.markdown(
            f'<div class="order-card">'
            f'<div style="display:flex;justify-content:space-between;align-items:flex-start;">'
            f'  <div>'
            f'    <div class="order-number">{order["order_number"]}</div>'
            f'    <div class="order-model">📋 {order["motor_model"]} — Cant: {order["quantity"]} u. · {order["supplier"]}</div>'
            f'  </div>'
            f'  <div style="text-align:right;">'
            f'    {_badge(order["status"])}'
            f'    <div class="act-meta" style="margin-top:4px;">Creado: {order["created_at"]}</div>'
            f'  </div>'
            f'</div>'
            f'<div style="margin-top:10px;">{_progress_bar(prog)}</div>'
            f'<div style="margin-top:10px;display:flex;align-items:center;gap:12px;">'
            f'  {_sem_dot(sem_color)}'
            f'  <span style="font-size:.80rem;color:#4B5563;">'
            f'    {"Actividad activa: <strong>" + active_act["name"] + "</strong>" if active_act else "Sin actividad activa"}'
            f'  </span>'
            f'  {"· Resp: " + _avatar(responsible.get("avatar","?"), responsible.get("color","#ccc")) + "<span style=\'font-size:.78rem;\'>" + responsible.get("name","") + "</span>" if responsible_key else ""}'
            f'</div>'
            f'</div>',
            unsafe_allow_html=True,
        )
        col_btn, _ = st.columns([1, 4])
        with col_btn:
            if st.button("Ver actividades →", key=f"open_order_{order['id']}"):
                st.session_state["selected_order_id"] = order["id"]
                st.session_state["page"] = "activities"
                st.rerun()


# ══════════════════════════════════════════════════════════
#  PÁGINA 3 — ACTIVIDADES DEL PEDIDO
# ══════════════════════════════════════════════════════════

def page_activities() -> None:
    data     = _app_data()
    user     = st.session_state["user"]
    order_id = st.session_state.get("selected_order_id")

    orders    = get_orders_for_user(data, user["username"])
    order_map = {o["order_number"]: o["id"] for o in orders}

    col_sel, col_back = st.columns([3, 1])
    with col_sel:
        selected_label = st.selectbox(
            "Seleccionar pedido",
            options=list(order_map.keys()),
            index=0 if not order_id else
                  next((i for i, k in enumerate(order_map) if order_map[k] == order_id), 0),
            key="order_selector",
        )
    with col_back:
        st.markdown("<br>", unsafe_allow_html=True)
        if st.button("← Tablero", use_container_width=True):
            st.session_state["page"] = "dashboard"
            st.rerun()

    if not selected_label:
        st.info("No hay pedidos disponibles.")
        return

    order_id = order_map[selected_label]
    st.session_state["selected_order_id"] = order_id
    order = next((o for o in data["orders"] if o["id"] == order_id), None)
    if not order:
        st.error("Pedido no encontrado.")
        return

    acts = order.get("activities", [])
    prog = order.get("progress", 0)

    sem_counts = {"green": 0, "yellow": 0, "red": 0, "gray": 0}
    for a in acts:
        c = get_semaphore(a)
        sem_counts[c] = sem_counts.get(c, 0) + 1

    st.markdown(
        f'<div style="background:#fff;border-radius:12px;padding:20px 24px;'
        f'margin-bottom:18px;box-shadow:0 1px 6px rgba(13,43,110,.10);border-top:4px solid #0D2B6E;">'
        f'<div style="display:flex;justify-content:space-between;flex-wrap:wrap;gap:8px;">'
        f'<div>'
        f'  <div style="font-family:\'Barlow Condensed\',sans-serif;font-size:1.5rem;font-weight:800;color:#0D2B6E;">'
        f'  📋 {order["order_number"]}</div>'
        f'  <div style="font-size:.85rem;color:#4B5563;margin-top:3px;">'
        f'  {order["motor_model"]} &nbsp;·&nbsp; {order["quantity"]} unidades &nbsp;·&nbsp; {order["supplier"]}</div>'
        f'</div>'
        f'<div style="text-align:right;">{_badge(order["status"])}'
        f'  <div style="font-size:.78rem;color:#8592A3;margin-top:4px;">Creado: {order["created_at"]}</div>'
        f'</div></div>'
        f'<div style="margin-top:14px;">{_progress_bar(prog)}</div>'
        f'<div style="margin-top:10px;display:flex;gap:14px;flex-wrap:wrap;">'
        f'  {_sem_dot("green")}<span style="font-size:.8rem;">{sem_counts["green"]} en tiempo</span>'
        f'  {_sem_dot("yellow")}<span style="font-size:.8rem;">{sem_counts["yellow"]} en riesgo</span>'
        f'  {_sem_dot("red")}<span style="font-size:.8rem;">{sem_counts["red"]} vencidas</span>'
        f'</div></div>',
        unsafe_allow_html=True,
    )

    # ── EXPORTAR ─────────────────────────────────────────────
    col_exp, _ = st.columns([1, 4])
    with col_exp:
        excel_bytes = _export_order_excel(order)
        if excel_bytes:
            st.download_button(
                label="⬇️  Exportar a Excel",
                data=excel_bytes,
                file_name=f"{order['order_number'].replace('/','-')}_actividades.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True,
            )
        else:
            st.button("⬇️  Exportar a Excel", disabled=True, use_container_width=True,
                      help="Instala openpyxl para habilitar: pip install openpyxl")

    st.markdown('<div class="section-header">🔢 Actividades del Proceso (19)</div>',
                unsafe_allow_html=True)

    current_phase = None
    for act in acts:
        phase = act.get("phase", "")
        if phase != current_phase:
            current_phase = phase
            phase_acts = [a for a in acts if a.get("phase") == phase]
            phase_done = sum(1 for a in phase_acts if a["status"] == "completed")
            st.markdown(
                f'<div style="margin:18px 0 8px;display:flex;align-items:center;gap:8px;">'
                f'<span style="font-family:\'Barlow Condensed\',sans-serif;font-size:1.1rem;'
                f'font-weight:700;color:#0D2B6E;">{phase}</span>'
                f'<span style="font-size:.72rem;color:#8592A3;">({phase_done}/{len(phase_acts)} completadas)</span>'
                f'</div>',
                unsafe_allow_html=True,
            )
        _render_activity_row(act, order, user, data)


def _render_activity_row(act: dict, order: dict, user: dict, data: dict) -> None:
    status          = act.get("status", "pending")
    sem_color       = get_semaphore(act)
    row_cls         = _act_row_class(status)
    responsible_key = act.get("responsible_key", "")
    responsible     = USERS.get(responsible_key, {})
    is_mine         = (responsible_key == user["username"])

    days_left  = _working_days_remaining(act.get("due_date","")) if act.get("due_date") else None
    days_html  = _days_badge(days_left) if status in ("in_progress","waiting_closure") else ""

    st.markdown(
        f'<div class="act-row {row_cls}">'
        f'<div style="display:flex;justify-content:space-between;align-items:flex-start;flex-wrap:wrap;">'
        f'  <div style="display:flex;align-items:center;gap:8px;">'
        f'    <span style="font-family:\'Barlow Condensed\',sans-serif;font-size:.85rem;'
        f'    font-weight:700;color:#8592A3;min-width:22px;">{act["id"]:02d}</span>'
        f'    {_sem_dot(sem_color)}'
        f'    <div><div class="act-name">{act["name"]}</div>'
        f'    <div class="act-meta">{act["description"]}</div></div>'
        f'  </div>'
        f'  <div style="text-align:right;">'
        f'    <span style="font-size:.75rem;padding:3px 10px;border-radius:20px;'
        f'    background:#F3F4F6;color:#374151;font-weight:600;">{STATUS_LABELS.get(status,"")}</span>'
        f'    <div class="act-meta" style="margin-top:4px;">'
        f'    Resp: {_avatar(responsible.get("avatar","?"), responsible.get("color","#ccc"))}'
        f'    {responsible.get("name","")}</div>'
        f'  </div>'
        f'</div>'
        f'<div class="act-meta" style="margin-top:6px;">'
        f'  Plazo: {act["days_allocated"]} días hábiles'
        f'  {"  ·  Inicio: " + act["start_date"] if act.get("start_date") else ""}'
        f'  {"  ·  Límite: " + act["due_date"] if act.get("due_date") else ""}'
        f'  {days_html}'
        f'  {"  ·  Cierre: " + act["completion_date"] if act.get("completion_date") else ""}'
        f'</div>'
        + (f'<div style="margin-top:4px;font-size:.75rem;color:#6B7280;">📎 Evidencia: {act["evidence_name"]}</div>'
           if act.get("evidence_name") else "")
        + '</div>',
        unsafe_allow_html=True,
    )

    # ── Galería de evidencias — visible para TODOS los usuarios ────────────
    _render_evidence_gallery(act)

    if is_mine and status == "in_progress":
        with st.expander(f"📋  Solicitar cierre — Actividad {act['id']:02d}: {act['name']}", expanded=False):
            with st.form(f"closure_form_{order['id']}_{act['id']}"):
                notes    = st.text_area("Observaciones / notas de cierre", height=90)
                evidence = st.file_uploader(
                    "📸  Adjuntar evidencia (imagen, PDF, Excel…)",
                    type=["pdf", "png", "jpg", "jpeg", "xlsx", "docx"],
                    key=f"ev_{order['id']}_{act['id']}",
                )
                submitted = st.form_submit_button("✅  Enviar solicitud de cierre",
                                                   type="primary", use_container_width=True)
                if submitted:
                    ev_name = ev_url = ev_thumb = ev_file_id = None

                    # ── Subir evidencia a Google Drive ───────────────────────
                    if evidence:
                        ev_bytes = evidence.read()
                        ev_name  = evidence.name
                        with st.spinner("📤 Subiendo evidencia a Google Drive…"):
                            ev_file_id, ev_url, ev_thumb = _upload_evidence_to_drive(
                                ev_bytes, ev_name,
                                order.get("order_number", ""),
                                act["id"],
                            )
                        if not ev_url:
                            st.warning(
                                "⚠️ La evidencia no pudo subirse a Drive "
                                "(el error exacto aparece arriba). "
                                "La actividad se cerrará de todas formas."
                            )

                    ok, msg = request_closure(
                        data, order["id"], act["id"],
                        user["username"], ev_name, None, notes,
                    )
                    if ok:
                        ok2, msg2, alerts = approve_closure(data, order["id"], act["id"])
                        if ok2:
                            # ── Guardar referencia Drive en la actividad ─────
                            for o in data["orders"]:
                                if o["id"] == order["id"]:
                                    for a in o.get("activities", []):
                                        if a["id"] == act["id"]:
                                            if ev_name and ev_url:
                                                if "evidences" not in a:
                                                    a["evidences"] = []
                                                a["evidences"].append({
                                                    "name":      ev_name,
                                                    "file_id":   ev_file_id or "",
                                                    "url":       ev_url,
                                                    "thumb_url": ev_thumb or "",
                                                    "user":      USERS.get(user["username"], {}).get("name", user["username"]),
                                                    "ts":        datetime.now().strftime("%d/%m/%Y %H:%M"),
                                                })
                                            _log_history(a, user["username"],
                                                         "Actividad cerrada",
                                                         notes or "Sin notas")
                                            break
                            for alert in alerts:
                                send_activation_email(**alert)
                            st.success(f"✅ Actividad cerrada. {msg2}"
                                       + (" · Evidencia guardada en Drive 📁" if ev_url else ""))
                            _app_save(data)
                            st.rerun()
                        else:
                            st.warning(msg2)
                    else:
                        st.error(msg)

    elif status == "waiting_closure":
        st.markdown(
            '<div style="font-size:.80rem;color:#92400E;font-weight:600;'
            'padding:4px 10px;background:#FEF3C7;border-radius:6px;display:inline-block;'
            'margin-top:4px;">⏳ Cierre solicitado — procesando…</div>',
            unsafe_allow_html=True,
        )

    # ── HISTORIAL DE CAMBIOS ─────────────────────────────
    history = act.get("history", [])
    if history:
        with st.expander(f"🕐  Historial  ({len(history)} entrada(s))", expanded=False):
            for entry in reversed(history):
                st.markdown(
                    f'<div style="padding:6px 10px;border-left:3px solid #0D2B6E;'
                    f'margin-bottom:6px;background:#F8FAFF;border-radius:0 6px 6px 0;">'
                    f'<span style="font-size:.72rem;color:#6B7280;">{entry["ts"]}</span>  '
                    f'<strong style="font-size:.80rem;color:#0D2B6E;">{entry["user"]}</strong>  '
                    f'<span style="font-size:.80rem;color:#374151;">— {entry["action"]}</span>'
                    + (f'<br><span style="font-size:.75rem;color:#6B7280;margin-left:4px;">{entry["detail"]}</span>'
                       if entry.get("detail") else "")
                    + '</div>',
                    unsafe_allow_html=True,
                )

    # ── COMENTARIOS ─────────────────────────────────────
    comments = act.get("comments", [])
    exp_label = f"💬  Comentarios  ({len(comments)})" if comments else "💬  Agregar comentario"
    with st.expander(exp_label, expanded=False):
        for cm in comments:
            st.markdown(
                f'<div style="padding:6px 10px;border-left:3px solid #C41E2E;'
                f'margin-bottom:6px;background:#FFF8F8;border-radius:0 6px 6px 0;">'
                f'<span style="font-size:.72rem;color:#6B7280;">{cm["ts"]}</span>  '
                f'<strong style="font-size:.80rem;color:#C41E2E;">{cm["user"]}</strong><br>'
                f'<span style="font-size:.82rem;color:#374151;">{cm["text"]}</span></div>',
                unsafe_allow_html=True,
            )
        with st.form(f"comment_form_{order['id']}_{act['id']}"):
            new_comment = st.text_area("Nuevo comentario", height=70, label_visibility="collapsed",
                                       placeholder="Escribe un comentario sobre esta actividad…")
            if st.form_submit_button("➕  Agregar comentario", use_container_width=True):
                if new_comment.strip():
                    # Agregar en data y guardar
                    for o in data["orders"]:
                        if o["id"] == order["id"]:
                            for a in o.get("activities", []):
                                if a["id"] == act["id"]:
                                    if "comments" not in a:
                                        a["comments"] = []
                                    a["comments"].append({
                                        "ts":   datetime.now().strftime("%d/%m/%Y %H:%M"),
                                        "user": USERS.get(user["username"], {}).get("name", user["username"]),
                                        "text": new_comment.strip(),
                                    })
                                    _log_history(a, user["username"], "Comentario agregado",
                                                 new_comment.strip()[:80])
                                    break
                    _app_save(data)
                    st.rerun()
                else:
                    st.warning("El comentario no puede estar vacío.")


# ══════════════════════════════════════════════════════════
#  PÁGINA 4 — NUEVO PEDIDO (solo David González)
# ══════════════════════════════════════════════════════════

def page_new_order() -> None:
    user = st.session_state["user"]
    if not user.get("can_create_orders"):
        st.error("⛔ Acceso restringido. Solo David González puede crear nuevos pedidos.")
        return

    data  = _app_data()
    year  = datetime.today().year
    count = len([o for o in data["orders"]
                 if o.get("year") == year and o.get("status") != "cancelled"])

    st.markdown('<div class="section-header">➕ Registrar Nuevo Pedido de Motores</div>',
                unsafe_allow_html=True)

    remaining = MAX_ANNUAL_ORDERS - count
    color_bar = "#22C55E" if remaining > 5 else "#F59E0B" if remaining > 2 else "#EF4444"
    st.markdown(
        f'<div style="background:#fff;border-radius:10px;padding:16px 20px;'
        f'margin-bottom:20px;box-shadow:0 1px 4px rgba(13,43,110,.08);'
        f'border-left:4px solid {color_bar};">'
        f'<span style="font-weight:700;color:{color_bar};">Pedidos {year}: {count}</span>'
        f'</div>',
        unsafe_allow_html=True,
    )

    if remaining == 0:
        st.error(f"🚫 Límite anual de {MAX_ANNUAL_ORDERS} pedidos alcanzado para {year}.")
        return

    with st.form("new_order_form"):
        col1, col2 = st.columns(2)
        with col1:
            motor_model  = st.text_input("Nombre del pedido",
                                         placeholder="Ej. Compra Motores 40HP — Temporada 2026")
            quantity     = st.number_input("Cantidad de motores", min_value=1, value=1, step=1)
        with col2:
            supplier     = st.text_input("Monto aproximado de O.C.",
                                         placeholder="Ej. $250,000 MXN")
            order_date   = st.date_input("Fecha de Pedido",
                                         value=datetime.today(),
                                         help="Puedes seleccionar una fecha anterior para registrar pedidos ya realizados.")

        notes = st.text_area("Justificación / Notas adicionales", height=100,
                             placeholder="Descripción del requerimiento, destino, prioridad…")

        st.markdown("<br>", unsafe_allow_html=True)
        col_b1, col_b2 = st.columns([1, 3])
        with col_b1:
            submitted = st.form_submit_button("🚀  Crear pedido", type="primary", use_container_width=True)
        with col_b2:
            st.info("Al crear el pedido se activará automáticamente la primera actividad "
                    "asignada a **David González** con fecha límite de 1 día hábil.")

        if submitted:
            if not motor_model.strip():
                st.error("El campo Nombre del pedido es obligatorio.")
            else:
                ok, msg, new_order = create_order(
                    data, motor_model, int(quantity),
                    supplier.strip(), notes, user["username"],
                    order_date.year,
                    order_date.strftime("%Y-%m-%d"),
                )
                if ok:
                    st.success(f"🎉 {msg}")
                    st.balloons()
                    _app_save(data)
                    st.session_state["selected_order_id"] = new_order["id"]
                    st.session_state["page"] = "activities"
                    st.rerun()
                else:
                    st.error(msg)



# ══════════════════════════════════════════════════════════
#  MAIN ROUTER
# ══════════════════════════════════════════════════════════

def main() -> None:
    inject_css()

    if "user" not in st.session_state:
        st.session_state["user"] = None
    if "page" not in st.session_state:
        st.session_state["page"] = "login"
    if "selected_order_id" not in st.session_state:
        st.session_state["selected_order_id"] = None

    if st.session_state["user"] is None:
        page_login()
        return

    _check_daily_alerts()

    page = render_sidebar()

    if page == "dashboard":
        page_dashboard()
    elif page == "activities":
        page_activities()
    elif page == "new_order":
        page_new_order()
    else:
        page_dashboard()


if __name__ == "__main__":
    main()
